import os
import pandas as pd
import traceback
from collections import defaultdict
import streamlit as st
from io import BytesIO
from zipfile import ZipFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ========== Utility Functions ==========
def safe_strip(x):
    if pd.isna(x):
        return ''
    return str(x).strip()

def check_clashes(subject_rolls):
    clashes = []
    subs = list(subject_rolls.keys())
    for i in range(len(subs)):
        for j in range(i + 1, len(subs)):
            s1, s2 = subs[i], subs[j]
            inter = subject_rolls[s1].intersection(subject_rolls[s2])
            if inter:
                for r in sorted(inter):
                    clashes.append((s1, s2, r))
    return clashes

def find_building_allocation(subject, count, rooms_by_building, rooms_info, eff_cap_func):
    allocation = []
    for b, room_list in rooms_by_building.items():
        total = sum(eff_cap_func(r) for r in room_list if rooms_info[r]['remaining'] > 0)
        if total >= count:
            for r in room_list:
                if count <= 0:
                    break
                rem = rooms_info[r]['remaining']
                if rem <= 0:
                    continue
                take = min(rem, count)
                allocation.append((r, take))
                count -= take
            if count == 0:
                return allocation
    for b, room_list in rooms_by_building.items():
        for r in room_list:
            if count <= 0:
                break
            rem = rooms_info[r]['remaining']
            if rem <= 0:
                continue
            take = min(rem, count)
            allocation.append((r, take))
            count -= take
        if count <= 0:
            return allocation
    return allocation

# ========== Core Logic ==========
def allocate_for_slot(date, day, slot, subjects, students_by_subject, rooms_df, buffer, density, roll_name_map):
    subject_rolls, subject_counts = {}, {}
    for s in subjects:
        rolls = set(safe_strip(r) for r in students_by_subject.get(s, []) if safe_strip(r))
        subject_rolls[s] = rolls
        subject_counts[s] = len(rolls)

    clashes = check_clashes(subject_rolls)
    clash_msgs = []
    if clashes:
        for s1, s2, r in clashes:
            clash_msgs.append(f"‚ö†Ô∏è Clash: {r} in {s1} and {s2}")

    rooms_info, rooms_by_building = {}, defaultdict(list)
    for _, row in rooms_df.iterrows():
        room = safe_strip(row['Room No.'])
        building = safe_strip(row['Block']) or 'UnknownBlock'
        cap = int(row['Exam Capacity'])
        rooms_info[room] = {'building': building, 'capacity': cap, 'remaining': 0}
        rooms_by_building[building].append(room)

    def eff_cap_room(r):
        return max(0, rooms_info[r]['capacity'] - int(buffer))

    def eff_cap_per_sub(r):
        ec = eff_cap_room(r)
        return ec // 2 if density == 'sparse' else ec

    for r in rooms_info:
        rooms_info[r]['remaining'] = eff_cap_per_sub(r)

    total_students = sum(subject_counts.values())
    total_avail = sum(eff_cap_per_sub(r) for r in rooms_info)
    if total_students > total_avail:
        clash_msgs.append("‚ùå Cannot allocate due to excess students")

    subj_assignments = {s: [] for s in subjects}
    for s in sorted(subjects, key=lambda x: -subject_counts[x]):
        count = subject_counts[s]
        if count == 0:
            continue
        alloc = find_building_allocation(s, count, rooms_by_building, rooms_info, eff_cap_per_sub)
        assigned = 0
        for r, take in alloc:
            rolls_to_take = list(subject_rolls[s])[assigned:assigned + take]
            subj_assignments[s].append({'room': r, 'rolls': rolls_to_take})
            assigned += len(rolls_to_take)
            rooms_info[r]['remaining'] -= len(rolls_to_take)

    # Prepare overall and seats_left tables
    overall_rows = []
    seats_left_rows = []
    for s in subjects:
        for p in subj_assignments[s]:
            r = p['room']
            rolls = p['rolls']
            overall_rows.append({
                'Date': date,
                'Day': day,
                'course_code': s,
                'Room': r,
                'Allocated_students_count': len(rolls),
                'Roll_list(semicolon separated)': ';'.join(rolls)
            })
    for r, info in rooms_info.items():
        allocated = info['capacity'] - info['remaining']
        seats_left_rows.append({
            'Room No.': r,
            'Exam Capacity': info['capacity'],
            'Block': info['building'],
            'Alloted': allocated,
            'Vacant': info['remaining']
        })

    return subj_assignments, overall_rows, seats_left_rows, clashes

# ========== Streamlit App ==========
st.title("ü™ë Exam Seating Arrangement Generator")
st.write("Upload Excel with sheets: **in_timetable**, **in_course_roll_mapping**, **in_roll_name_mapping**, **in_room_capacity**")

uploaded_file = st.file_uploader("Upload input_data.xlsx", type=["xlsx"])
buffer = st.number_input("Enter Buffer", min_value=0, value=5)
density = st.radio("Select Arrangement Type", ["dense", "sparse"])

if uploaded_file:
    try:
        xls = pd.ExcelFile(uploaded_file)

        timetable = pd.read_excel(xls, 'in_timetable', dtype=str)
        students_df = pd.read_excel(xls, 'in_course_roll_mapping', dtype=str)
        rooms_df = pd.read_excel(xls, 'in_room_capacity', dtype=str)
        mapping_df = pd.read_excel(xls, 'in_roll_name_mapping', dtype=str)

        students_by_subject = defaultdict(list)
        for _, row in students_df.iterrows():
            s, r = safe_strip(row.get('course_code', '')), safe_strip(row.get('rollno', ''))
            if s and r:
                students_by_subject[s].append(r)

        roll_name_map = {safe_strip(r['Roll']): safe_strip(r['Name']) or 'Unknown Name'
                         for _, r in mapping_df.iterrows() if safe_strip(r['Roll'])}

        all_clashes = []
        overall_all = []
        seats_left_all = []
        zip_buffer = BytesIO()
        with ZipFile(zip_buffer, "w") as zipf:
            for _, trow in timetable.iterrows():
                date_val = trow['Date']
                if pd.isna(date_val):
                    date = ''
                elif isinstance(date_val, pd.Timestamp):
                    date = date_val.strftime('%Y-%m-%d')
                else:
                    date = str(date_val).split(' ')[0]
                date = safe_strip(date)

                day = safe_strip(trow['Day'])
                for slot_col in ['Morning', 'Evening']:
                    subjects = [safe_strip(s) for s in safe_strip(trow.get(slot_col, '')).split(';') if safe_strip(s)]
                    if not subjects:
                        continue
                    subj_assignments, overall_rows, seats_left_rows, clashes = allocate_for_slot(
                        date, day, slot_col, subjects, students_by_subject, rooms_df, buffer, density, roll_name_map)
                    all_clashes.extend(clashes)
                    overall_all.extend(overall_rows)
                    seats_left_all.extend(seats_left_rows)

                    # ZIP folder structure per day/slot/course/room
                    day_folder = f"{date}/{slot_col}"
                    for course, allocations in subj_assignments.items():
                        for alloc in allocations:
                            room = alloc['room']
                            rolls = alloc['rolls']
                            wb = Workbook()
                            ws = wb.active
                            ws.title = f"{course}_{room}"

                            # Top row info
                            ws.append([f"Course: {course}", f"Room: {room}", f"Date: {date}", f"Session: {slot_col}"])
                            for cell in ws[1]:
                                cell.font = Font(bold=True)
                            ws.append([])  # empty row

                            # Column headers
                            ws.append(['Roll', 'Student Name', 'Signature'])
                            for r in rolls:
                                ws.append([r, roll_name_map.get(r, 'Unknown Name'), ''])

                            # Bottom 10 rows for TAs / Invigilators
                            for ta in ['TA1','TA2','TA3','TA4','TA5','Invigilator1','Invigilator2','Invigilator3','Invigilator4','Invigilator5']:
                                ws.append([ta,'',''])

                            # Adjust column width and center
                            for col in ws.columns:
                                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                                col_letter = get_column_letter(col[0].column)
                                ws.column_dimensions[col_letter].width = max(15, max_length + 2)
                                for cell in col:
                                    cell.alignment = Alignment(horizontal='center', vertical='center')

                            # Save sheet to buffer
                            excel_buffer = BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            filename = f"{day_folder}/{date}_{course}_{room}_{slot_col}.xlsx"
                            zipf.writestr(filename, excel_buffer.read())

        zip_buffer.seek(0)
        st.success("‚úÖ Seating arrangement generated successfully!")

        # Preview tables
        st.subheader("Overall Seating Arrangement Preview")
        st.dataframe(pd.DataFrame(overall_all).head(10))

        st.subheader("Seats Left Preview")
        st.dataframe(pd.DataFrame(seats_left_all).head(10))

        if all_clashes:
            st.warning("Some clashes detected:")
            for c in all_clashes:
                st.text(c)

        # Download ZIP
        st.download_button(
            label="üì• Download Results (Structured ZIP)",
            data=zip_buffer,
            file_name="seating_arrangement_structured.zip",
            mime="application/zip"
        )

        # Also allow download of consolidated overall/seats_left Excel
        with BytesIO() as xls_buf:
            with pd.ExcelWriter(xls_buf, engine='xlsxwriter') as writer:
                pd.DataFrame(overall_all).to_excel(writer, index=False, sheet_name='Overall')
                pd.DataFrame(seats_left_all).to_excel(writer, index=False, sheet_name='Seats_Left')
            xls_buf.seek(0)
            st.download_button(
                label="üì• Download Consolidated Excel",
                data=xls_buf,
                file_name="op_overall_seating_arrangement.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error processing file: {e}")
        st.text(traceback.format_exc())
